from flask import Flask, render_template, request, redirect, url_for, jsonify, flash
import os
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import distinct
from filters import enumerate_filter
import variables
from datetime import datetime , timedelta
from collections import defaultdict
app = Flask(__name__)
ALLOWED_EXTENSIONS = {'xlsx'}
hjdia='12'
hjmes='09'
hjano='2023'
app.secret_key = 'secret-key'

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in \
        ALLOWED_EXTENSIONS


app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///data.db'
db = SQLAlchemy(app)


class Data(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    time = db.Column(db.Time, nullable=False)
    soil_humidity = db.Column(db.Float, nullable=False)
    ambient_humidity = db.Column(db.Float, nullable=False)
    ambient_temperature = db.Column(db.Float, nullable=False)
    water_volume = db.Column(db.Float, nullable=False)

# tabela para armazenar a senha
class Code(db.Model): 
    id = db.Column(db.Integer, primary_key=True)
    password = db.Column(db.String, nullable=False)


with app.app_context():
    db.create_all()

# Registrando o filtro criado em filters.py
app.jinja_env.filters['enumerate'] = enumerate_filter


@app.route("/")
def index():
    mes = datetime.now().strftime('%h')
    overview_list = zip(variables.svg_overview_list, variables.overview_desc)

    # Realizando um select * em nossa tabela e guardando estes valores em
    # uma variável
    dados_db = Data.query.all()

    # Aqui estou pegando todos os valores do campo date e os tornando único,
    # pois há vários dias iguais gravados no banco de dados
    datas = db.session.query(distinct(Data.date)).all()

    datas_index = 0  # variável necessária para atribuir ao atributo key a data completa

    return render_template(
        'index.html',
        overview_list=overview_list,
        dados_db=dados_db,
        datas=datas,
        datas_index=datas_index,
        mes=mes)


@app.route("/trocar", methods=["POST"])
def trocar():
    with app.app_context():
        codigo = Code.query.first()
        senha = request.form["senha"]
        senhanv = request.form['senhanv']

        # Validação dos campos
        if not senha or not senhanv:
            return redirect(url_for("delete_data", error="Por favor, preencha todos os campos."))

        if codigo.password == senha:
            if len(senhanv) < 6:  
                return redirect(url_for("delete_data", error="O novo código deve ter pelo menos 6 caracteres."))

            # Atualiza o código para o novo código fornecido
            codigo.password = senhanv
            db.session.commit()
            return redirect(url_for("delete_data", success="Código atualizado com sucesso."))
    return redirect(url_for("delete_data", error="Senha incorreta."))


@app.route("/show-data")
def show_data():
    data = request.args.get('date')
    data = data.split()
    ano = data[0].replace('datetime.date', '').replace('(', '').replace(',', '')
    mes = data[1].replace(',', '')
    dia = data[2].replace(')', '').replace(',', '')

    dia_swiper = dia
    if len(mes) < 2:
        mes = f'0{mes}'

    if len(dia) < 2:
        dia = f'0{dia}'

    date_key = f'{ano}-{mes}-{dia}'

    all_dates = Data.query.filter_by(date=f'{ano}-{mes}-{dia}').all()
    return render_template('show_data.html', all_dates=all_dates, date_key=date_key, dia_swiper=dia_swiper)


@app.route("/add-data")
def add_data():
    return render_template('add_data.html')


@app.route("/delete-data")
def delete_data():
    data = request.args.get('date')
    if data:
        data = data.split()
        ano = data[0].replace('datetime.date', '').replace('(', '').replace(',', '')
        mes = data[1].replace(',', '')
        dia = data[2].replace(')', '').replace(',', '')

        # Valor que irá aparecer na bolinha vermelha
        dia_swiper = dia
        if len(mes) < 1:
            mes = f'0{mes}'

        if len(dia) < 2:
            dia = f'0{dia}'
        # Valor que será salvo no atributo key   
        date_key = f'{ano}-{mes}-{dia}'
    else:
        date_key = 'no value'
        dia_swiper = 'x'

    error_msg = request.args.get('error')
    if error_msg:
        error = error_msg
    else: error = ''

    success_msg = request.args.get('success')
    if success_msg:
        success = success_msg
    else: success = ''

    # Função que coloca a chave padrão no DB se ele estiver vazio
    with app.app_context():
        empty_check = Code.query.count()
        if empty_check == 0:
            start_pwrd = Code(id=1, password='c0d1g0')
            db.session.add(start_pwrd)
            db.session.commit()

    return render_template('delete_data.html', date_key=date_key, dia_swiper=dia_swiper, success=success, error=error)


@app.route("/del-dia", methods=["POST"])
def del_dia():
    data = request.args.get('date')
    data = datetime.strptime(data, "%Y-%m-%d").date()
    print(data)

    senha = request.form['senha']
    with app.app_context():
        codigo = Code.query.first()
    if senha == codigo.password:
        db.session.query(Data).filter_by(date=data).delete()
        db.session.commit()
        return render_template("delete_data.html", success="Dia deletado com sucesso.")
    return render_template("delete_data.html", error="Código de segurança errado.")


@app.route("/statistics", methods=["GET", "POST"])
def statistics():
    lista = []
    msg = ''

    # Obter todas as datas disponíveis no banco de dados
    available_dates = [str(date[0]) for date in db.session.query(distinct(Data.date)).all()]

    if request.method == "POST":
        interval = request.form.get("interval")
        if interval:
            interval = int(interval)
            end_date = datetime.strptime(available_dates[-1], '%Y-%m-%d')
            start_date = end_date - timedelta(days=interval)
            lista = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range((end_date - start_date).days + 1)]
        else:
            date = str(request.form["data"]).split("-")
            dia = int(date[2])
            mes = int(date[1])
            ano = int(date[0])
            datef = str(request.form["dataf"]).split("-")
            diaf = int(datef[2])
            mesf = int(datef[1])
            anof = int(datef[0])

            start_date = datetime(ano, mes, dia)
            end_date = datetime(anof, mesf, diaf)
            
            current_date = start_date
            while current_date <= end_date:
                lista.append(current_date.strftime('%Y-%m-%d'))
                current_date += timedelta(days=1)
        
        print(lista)
    
    listasoilh = []
    listaambienth = []
    listaambientt = []
    listawater = []
    
    for date_str in lista:
        try:
            dia = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            msg = "Data inválida"
    
        dados_por_dia = db.session.query(Data).filter_by(date=dia).all()
        if not dados_por_dia:
            msg = "Não há dados para a data selecionada"
            return render_template('statistics.html', msg=msg, available_dates=available_dates)

        # Dicionário para armazenar somas e contagens por atributo
        soma_por_atributo = defaultdict(float)
        contagem_por_atributo = defaultdict(int)

        # Calcular a soma e contagem para cada atributo
        for data in dados_por_dia:  
            soma_por_atributo["soil_humidity"] += round(data.soil_humidity, 2)
            soma_por_atributo["ambient_humidity"] += round(data.ambient_humidity, 2)
            soma_por_atributo["ambient_temperature"] += round(data.ambient_temperature, 2)
            soma_por_atributo["water_volume"] += round(data.water_volume, 2)
            contagem_por_atributo["soil_humidity"] += 1
            contagem_por_atributo["ambient_humidity"] += 1
            contagem_por_atributo["ambient_temperature"] += 1
            contagem_por_atributo["water_volume"] += 1
                
        # Calcular as médias
        media_por_atributo = {atributo: round(soma_por_atributo[atributo] / contagem_por_atributo[atributo], 1) for atributo in soma_por_atributo}
        listasoilh.append(media_por_atributo["soil_humidity"])
        listaambienth.append(media_por_atributo["ambient_humidity"])
        listaambientt.append(media_por_atributo["ambient_temperature"])
        listawater.append(media_por_atributo["water_volume"])
       
    print("Lista Soil Humidity:", listasoilh)
    print("Lista Ambient Humidity:", listaambienth)
    print("Lista Ambient Temperature:", listaambientt)
    print("Lista Water Volume:", listawater)
    print(lista)
   
    return render_template('statistics.html', 
                           listasoilh=listasoilh,
                           listaambienth=listaambienth,
                           listaambientt=listaambientt,
                           listawater=listawater,
                           lista=lista,
                           msg=msg,
                           available_dates=available_dates)


@app.route("/upload", methods=["POST"])
@app.route("/upload", methods=["POST"])
def upload():
    if request.method == "POST":
        file = request.files["file"]
        if not os.path.exists('./uploads'):
            os.mkdir('./uploads')

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join("uploads", filename))

            wb = load_workbook(os.path.join("uploads", filename))
            sheet = wb.active
            for row in range(1, sheet.max_row + 1):
                try:
                    info = Data()
                    info.date = datetime.strptime(str(sheet.cell(row=row, column=2).value).split(" ")[0], "%Y-%m-%d")
                    info.time = sheet.cell(row=row, column=3).value
                    info.soil_humidity = sheet.cell(row=row, column=4).value
                    info.ambient_humidity = sheet.cell(row=row, column=5).value
                    info.ambient_temperature = sheet.cell(row=row, column=6).value
                    info.water_volume = sheet.cell(row=row, column=7).value
                    db.session.add(info)
                except:
                    continue
            db.session.commit()
            flash("Os dados foram enviados com sucesso", "success")
        else:
            flash("Arquivo inválido", "error")

        return redirect(url_for('add_data'))


@app.route('/get_data')
def get_data():
    dia = request.args.get('dia')
    dados_dia = db.session.query(Data).filter_by(date=dia).all()

    res_dados_dia = [
        {
            "soil_humidity": data.soil_humidity,
            "ambient_humidity": data.ambient_humidity,
            "ambient_temperature": data.ambient_temperature,
            "water_volume": data.water_volume
        }
        for data in dados_dia
    ]

    return jsonify(res_dados_dia)
